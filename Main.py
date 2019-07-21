import PIL.Image as Image_1
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Color,PatternFill
from Tkinter import *
import tkFileDialog, Tkconstants
root=Tk()
root.withdraw()
fname= tkFileDialog.askopenfilename()


workbook=Workbook()
ws=workbook.create_sheet(title="tree")
#ws.title="car"
im=Image_1.open(fname)
(w,h)=im.size
(w,h)=(512,512*1.09)
im=im.resize((512,512),Image_1.ANTIALIAS)
imagesize=im.size

pix= im.load()
def htmlcolor(i,j):
    x,y,z=pix[i,j]
    x,y,z=hex(x),hex(y),hex(z)
    x,y,z=x[2:],y[2:],z[2:]
    if len(x)<2:
        x='0'+x
    if len(y)<2:
        y='0'+y
    if(len(z)<2):
        x='0'+x
    
    htmlcolor='00'+x+y+z
    return htmlcolor
for i in range(imagesize[0]):
    for j in range(imagesize[1]):
        colore=htmlcolor(i,j)
        fill=PatternFill("solid",fgColor=colore)
        cell=ws.cell(row=i+1,column=j+1)
        cell.fill=fill
        
        #co='00FF00FF'
        #print(colore,"___",i,j,"_____",pix[i,j])
        #colore= openpyxl.styles.colors.Color(rgb=htmlcolor(i,j))
        #co= openpyxl.styles.colors.Color(rgb='00FF0000')
        #fill=openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=co)
    
#nefname=tkFileDialog.askopenfilename()
nefname=raw_input('Enter the Xcel Sheetname')

nefname=str(nefname+'.xlsx')
workbook.save(nefname)
